# Automate Occultation recordings using SharpCap - GUI Only Version
# Retrieves user occultations events from Occult Watcher Cloud
# Maintains a list of the latest events from Occult Watcher Cloud
# and a set of events that have been retrieved that are less than 14 days old
## Key Features Added:

# ### **Exposure Editing Functionality:**
# 1. **Double-click exposure cell** to edit exposure for any event
# 2. **Edit Exposure button** in toolbar for selected event
# 3. **ExposureEditDialog** with:
#    - Current exposure display (shows if custom or calculated)
#    - Quick exposure buttons (10ms, 20ms, 40ms, etc.)
#    - Reset to calculated exposure option
#    - Validation for reasonable exposure ranges (1-10000ms)

# ### **Automatic Sequence Regeneration:**
# 1. **Immediate regeneration** after exposure editing
# 2. **Template selection** for regeneration
# 3. **Single event sequence generation** method
# 4. **Progress feedback** during sequence creation

# ### **Enhanced Event Management:**
# 1. **Custom exposure tracking** with visual indicators (asterisk *)
# 2. **Exposure persistence** throughout session
# 3. **Grid updates** to reflect exposure changes
# 4. **Detailed event information** showing custom vs calculated exposure

# ### **User Experience Improvements:**
# 1. **Visual indicators** for custom exposures in the grid
# 2. **Confirmation dialogs** for sequence regeneration
# 3. **Error handling** with user-friendly messages
# 4. **Status updates** during operations

# ### **Usage Instructions:**
# 1. **Edit Single Exposure:** Double-click the exposure cell or select event and click "Edit Exposure"
# 2. **Quick Settings:** Use preset exposure buttons for common values
# 3. **Auto-Regenerate:** Choose "Yes" when prompted to regenerate sequence after editing
# 4. **Visual Feedback:** Custom exposures show with "*" in the grid
# 5. **Reset Option:** Use "Reset to Calculated" to restore automatic exposure calculation

# The application now provides a complete GUI-only solution with sophisticated exposure editing capabilities and automatic sequence regeneration, making it easy to customize individual event parameters while maintaining the automated workflow.
import os
import json
import binascii
import base64
import urllib.request
import math
from datetime import datetime, timedelta
import threading

# GUI Components
import clr
clr.AddReference("System.Windows.Forms")
clr.AddReference("System.Drawing")

from System.Drawing import Point, Size, SystemColors, Color, Font, FontStyle
from System.Windows.Forms import *
import System

###############################
# Enhanced Configuration Manager
###############################

class ConfigManager:
    """Manages all configuration and settings with persistent storage"""
    
    CONFIG_FILENAME = 'occultation_config.json'
    
    def __init__(self, config_folder=None):
        # Default configuration values
        self.default_config = {
            # User credentials
            'owc_user_email': 'your_owc_email',
            'owc_user_password': 'your_owc_password',
            
            # File paths
            'my_file_folder': os.path.normpath(r'C:\Users\AstroPC\Documents\SharpCap'),
            'my_occultations_file': 'occultations.json',
            'my_latest_occultations_file': 'occultations_latest.json',
            'sequence_path': '',  # Will be set to my_file_folder if empty
            
            # Recording parameters
            'base_duration': 60,
            'goto_lead_time': 240,
            'mag_for_40ms_exposure': 12.0,
            
            # Observer information for NA Report Form
            'observer_name': '',
            'observer_email': '',
            'observer_latitude': 0.0,
            'observer_longitude': 0.0,
            'observer_elevation': 0.0,
            
            # Telescope information for NA Report Form
            'telescope_aperture': 0,  # mm
            'telescope_focal_length': 0,  # mm
            'telescope_type': 'SCT including Cass and Mak',
            
            # API configuration
            'host': 'https://www.occultwatcher.net:443',
            'url_path': '/api2/v1/events/details-list',
            'apiKey': 'get and api key for your user from OWC',
            'URL_OCCELMNT_ENDPOINT_PATH': '/api2/v1/owc/event/my/%s/occelmnts'
        }
        
        # Set config folder
        if config_folder:
            self.config_folder = os.path.normpath(config_folder)
        else:
            # Use default folder or current directory if not accessible
            try:
                self.config_folder = os.path.normpath(self.default_config['my_file_folder'])
                if not os.path.exists(self.config_folder):
                    os.makedirs(self.config_folder, exist_ok=True)
            except:
                self.config_folder = os.path.normpath(os.getcwd())
        
        # Initialize configuration
        self.config = self.default_config.copy()
        self.load_config()
        
        # Set sequence_path to my_file_folder if empty
        if not self.config['sequence_path']:
            self.config['sequence_path'] = self.config['my_file_folder']
            
        # Ensure all paths use proper separators
        self._normalize_paths()
        
        # Change to working directory
        try:
            os.chdir(self.get_file_folder())
        except:
            print(f"Warning: Could not change to directory {self.get_file_folder()}")
    
    def _normalize_paths(self):
        """Normalize all path configurations"""
        path_keys = ['my_file_folder', 'sequence_path']
        for key in path_keys:
            if key in self.config and self.config[key]:
                self.config[key] = os.path.normpath(self.config[key])
    
    def get_config_path(self):
        """Get the full path to the configuration file"""
        return os.path.join(self.config_folder, self.CONFIG_FILENAME)
    
    def load_config(self):
        """Load configuration from file"""
        config_path = self.get_config_path()
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    saved_config = json.load(f)
                
                # Update current config with saved values
                for key, value in saved_config.items():
                    if key in self.default_config:
                        self.config[key] = value
                
                print(f"Configuration loaded from: {config_path}")
            else:
                print(f"No configuration file found. Using defaults.")
                self.save_config()  # Save defaults
        except Exception as e:
            print(f"Error loading configuration: {e}")
            print("Using default configuration")
    
    def save_config(self):
        """Save current configuration to file"""
        config_path = self.get_config_path()
        try:
            # Ensure config directory exists
            os.makedirs(self.config_folder, exist_ok=True)
            
            with open(config_path, 'w') as f:
                json.dump(self.config, f, indent=2)
            
            print(f"Configuration saved to: {config_path}")
            return True
        except Exception as e:
            print(f"Error saving configuration: {e}")
            return False
    
    def reset_to_defaults(self):
        """Reset configuration to default values"""
        self.config = self.default_config.copy()
        self._normalize_paths()
        return self.save_config()
    
    # User credentials
    def get_owc_email(self):
        return self.config['owc_user_email']
    
    def set_owc_email(self, email):
        self.config['owc_user_email'] = email
    
    def get_owc_password(self):
        return self.config['owc_user_password']
    
    def set_owc_password(self, password):
        self.config['owc_user_password'] = password
    
    # File paths
    def get_file_folder(self):
        return os.path.normpath(self.config['my_file_folder'])
    
    def set_file_folder(self, folder):
        self.config['my_file_folder'] = os.path.normpath(folder)
    
    def get_occultations_file(self):
        return self.config['my_occultations_file']
    
    def set_occultations_file(self, filename):
        self.config['my_occultations_file'] = filename
    
    def get_latest_occultations_file(self):
        return self.config['my_latest_occultations_file']
    
    def set_latest_occultations_file(self, filename):
        self.config['my_latest_occultations_file'] = filename
    
    def get_sequence_path(self):
        return os.path.normpath(self.config['sequence_path'])
    
    def set_sequence_path(self, path):
        self.config['sequence_path'] = os.path.normpath(path)
    
    # Recording parameters
    def get_base_duration(self):
        return self.config['base_duration']
    
    def set_base_duration(self, duration):
        self.config['base_duration'] = int(duration)
    
    def get_goto_lead_time(self):
        return self.config['goto_lead_time']
    
    def set_goto_lead_time(self, time):
        self.config['goto_lead_time'] = int(time)
    
    def get_mag_for_40ms_exposure(self):
        return self.config['mag_for_40ms_exposure']
    
    def set_mag_for_40ms_exposure(self, magnitude):
        self.config['mag_for_40ms_exposure'] = float(magnitude)
    
    # Observer configuration
    def get_observer_name(self):
        return self.config['observer_name']
    
    def set_observer_name(self, name):
        self.config['observer_name'] = name
    
    def get_observer_email(self):
        return self.config['observer_email']
    
    def set_observer_email(self, email):
        self.config['observer_email'] = email
    
    def get_observer_latitude(self):
        return self.config['observer_latitude']
    
    def set_observer_latitude(self, lat):
        self.config['observer_latitude'] = float(lat)
    
    def get_observer_longitude(self):
        return self.config['observer_longitude']
    
    def set_observer_longitude(self, lon):
        self.config['observer_longitude'] = float(lon)
    
    def get_observer_elevation(self):
        return self.config['observer_elevation']
    
    def set_observer_elevation(self, elev):
        self.config['observer_elevation'] = float(elev)
    
    # Telescope configuration
    def get_telescope_aperture(self):
        return self.config['telescope_aperture']
    
    def set_telescope_aperture(self, aperture):
        self.config['telescope_aperture'] = int(aperture)
    
    def get_telescope_focal_length(self):
        return self.config['telescope_focal_length']
    
    def set_telescope_focal_length(self, focal_length):
        self.config['telescope_focal_length'] = int(focal_length)
    
    def get_telescope_type(self):
        return self.config['telescope_type']
    
    def set_telescope_type(self, tel_type):
        self.config['telescope_type'] = tel_type
    
    # API configuration
    def get_host(self):
        return self.config['host']
    
    def set_host(self, host):
        self.config['host'] = host
    
    def get_api_key(self):
        return self.config['apiKey']
    
    def set_api_key(self, key):
        self.config['apiKey'] = key
    
    def get_full_url(self):
        """Get the complete API URL with key"""
        base_url = self.config['host'] + self.config['url_path']
        sep = '&' if '?' in base_url else '?'
        return base_url + sep + 'apikey=%s' % binascii.unhexlify(self.config['apiKey'].encode()).decode('ascii')
    
    def get_occelmnt_url(self):
        """Get the complete occelmnt URL with key"""
        base_url = self.config['host'] + self.config['URL_OCCELMNT_ENDPOINT_PATH']
        sep = '&' if '?' in base_url else '?'
        return base_url + sep + 'apikey=%s' % binascii.unhexlify(self.config['apiKey'].encode()).decode('ascii')
    
    def get_full_file_path(self, filename):
        """Get full path for a file in the configured folder"""
        return os.path.join(self.get_file_folder(), filename)
    
    def validate_config(self):
        """Validate current configuration"""
        errors = []
        
        # Check required fields
        if not self.config['owc_user_email']:
            errors.append("OWC email is required")
        
        if not self.config['owc_user_password']:
            errors.append("OWC password is required")
        
        # Check paths exist or can be created
        try:
            folder = self.get_file_folder()
            if not os.path.exists(folder):
                os.makedirs(folder, exist_ok=True)
        except Exception as e:
            errors.append(f"Cannot access/create file folder: {e}")
        
        try:
            seq_path = self.get_sequence_path()
            if not os.path.exists(seq_path):
                os.makedirs(seq_path, exist_ok=True)
        except Exception as e:
            errors.append(f"Cannot access/create sequence path: {e}")
        
        # Check numeric values
        if self.config['base_duration'] <= 0:
            errors.append("Base duration must be positive")
        
        if self.config['goto_lead_time'] <= 0:
            errors.append("GOTO lead time must be positive")
        
        if self.config['mag_for_40ms_exposure'] <= 0:
            errors.append("Magnitude for 40ms exposure must be positive")
        
        return errors

# Create global configuration instance
config = ConfigManager()

###############################
# Core Classes
###############################

class EventProcessor:
    """Handles event processing operations"""
    
    @staticmethod
    def load_occultations(filename):
        """Load occultations from JSON file"""
        try:
            full_path = config.get_full_file_path(filename)
            if os.path.exists(full_path):
                with open(full_path, 'r') as f:
                    return json.load(f)
        except Exception as ex:
            print(f"Error loading occultations: {ex}")
        return []
    
    @staticmethod
    def save_occultations(events_data, filename):
        """Save occultations to JSON file"""
        try:
            full_path = config.get_full_file_path(filename)
            with open(full_path, 'w') as f:
                json.dump(events_data, f, indent=2)
            return True
        except Exception as ex:
            print(f"Error saving occultations: {ex}")
            return False
    
    @staticmethod
    def merge_occultation_lists(existing, new, id_key='unique_id', retention_days=14):
        """Merge two lists of occultation dictionaries"""
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        merged_dict = {}
        
        # Add existing occultations
        for occ in existing:
            if occ is None:
                continue
            if datetime.strptime(occ['event_time'].split('T')[0], '%Y-%m-%d') > cutoff_date: 
                merged_dict[occ[id_key]] = occ
        
        # Add/update with new occultations
        added = 0
        updated = 0
        for occ in new:
            if occ is None:
                continue
            if occ[id_key] in merged_dict and datetime.strptime(occ['event_time'].split('T')[0], '%Y-%m-%d') > cutoff_date:
                updated += 1
                merged_dict[occ[id_key]] = occ
            if occ[id_key] not in merged_dict and datetime.strptime(occ['event_time'].split('T')[0], '%Y-%m-%d') > cutoff_date: 
                added += 1
                merged_dict[occ[id_key]] = occ
        
        result = list(merged_dict.values())
        print("Merge completed: {} existing + {} new = {} total ({} added, {} updated)".format(
            len(existing), len(new), len(result), added, updated))
        
        return result
    
    @staticmethod    
    def get_owc_events(url, username, password, data=None):
        """Get events from OW Cloud API"""
        credentials = f"{username}:{password}"
        encoded_credentials = base64.b64encode(credentials.encode('utf-8')).decode('utf-8')
        
        request = urllib.request.Request(url)
        request.add_header("Authorization", f"Basic {encoded_credentials}")
        request.add_header("Content-Type", "application/json")
        
        if data:
            request.data = json.dumps(data).encode('utf-8')
        
        response = urllib.request.urlopen(request)
        return json.loads(response.read().decode('utf-8'))

    @staticmethod
    def update_ow_cloud_events():
        """Get all your OWC announced events using configuration"""
        try:
            result = EventProcessor.get_owc_events(
                config.get_full_url(), 
                config.get_owc_email(), 
                config.get_owc_password()
            )
        except urllib.error.HTTPError as e:
            print(f"HTTP Error: {e.code} - {e.reason}")
            return []

        result = EventProcessor.process_owc_events(result, sitefilter='')
        EventProcessor.save_occultations(result, config.get_latest_occultations_file())
        latest_occultations = EventProcessor.load_occultations(config.get_latest_occultations_file())

        # Create new master if doesn't exist
        master_file = config.get_full_file_path(config.get_occultations_file())
        if not os.path.exists(master_file):
            print("File {} not found - creating new master occultations file".format(config.get_occultations_file()))
            EventProcessor.save_occultations(result, config.get_occultations_file())
        existing_occultations = EventProcessor.load_occultations(config.get_occultations_file())

        merged_occultations = EventProcessor.merge_occultation_lists(existing_occultations, latest_occultations, id_key='id', retention_days=14)
        EventProcessor.save_occultations(merged_occultations, config.get_occultations_file())
        return latest_occultations
    
    @staticmethod    
    def process_owc_events(owevents, sitefilter):
        """Process OWC events to extract the parameters"""
        occultations = []
        for owevent in owevents:
            name = owevent['Object']
            eventDuration = float(owevent['MaxDurSec'])
            eventId = owevent['Id']
            star_id = owevent['StarName']
            ra = float(owevent['RAJ2000Hours'])
            dec = float(owevent['DEJ2000Deg'])

            for station in owevent['Stations']:
                if station['IsOwnStation']:
                    eventTime = station['EventTimeUtc']
                    eventUncertainty = station['ErrorInTimeSec']
                    stationName = station['StationName']
                    latitude = station['Latitude']
                    longitude = station['Longitude']
                    starAz = station['StarAz']
                    starAlt = station['StarAlt']
                    starMag = owevent['StarMag']
                    combMag = station['CombMag']
                    magDrop = owevent['MagDrop']

                    if sitefilter != '' and not stationName.startswith(sitefilter):
                        print('Ignoring: %s due to site filter: %s' % (stationName, sitefilter))
                        continue

                    # Recording duration using config values
                    base_dur = config.get_base_duration()
                    recording_duration = round(base_dur + (eventDuration if eventDuration > 5 else 0) + 6*(eventUncertainty if eventUncertainty > 2 else 0))

                    # Calculate the Start/End Times
                    eventCenterTime = datetime.strptime(eventTime.split('.')[0], '%Y-%m-%dT%H:%M:%S')
                    startTime = eventCenterTime - timedelta(seconds=recording_duration/2.0)
                    endTime = eventCenterTime + timedelta(seconds=recording_duration/2.0)
                    gotoTime = eventCenterTime - timedelta(seconds=recording_duration/2.0 + config.get_goto_lead_time())
                    
                    eventCenterTime = eventCenterTime.strftime("%Y-%m-%dT%H:%M:%S")
                    startTime = startTime.strftime("%Y-%m-%dT%H:%M:%S")
                    endTime = endTime.strftime("%Y-%m-%dT%H:%M:%S")
                    gotoTime = gotoTime.strftime("%Y-%m-%dT%H:%M:%S")

                    # Get the occelmnt
                    occelmntUrl = config.get_occelmnt_url() % eventId
                    try:
                        eventOccelmnt = EventProcessor.get_owc_events(
                            occelmntUrl, 
                            config.get_owc_email(), 
                            config.get_owc_password()
                        )
                    except urllib.error.HTTPError as e:
                        print(f"HTTP Error: {e.code} - {e.reason}")
                        eventOccelmnt = None

                    if eventOccelmnt:
                        elements = eventOccelmnt['Occultations']['Event']['Elements'].split(',')
                        star = eventOccelmnt['Occultations']['Event']['Star'].split(',')
                        object_data = eventOccelmnt['Occultations']['Event']['Object'].split(',')
                        owcloudurl = 'https://cloud.occultwatcher.net' + eventOccelmnt['Occultations']['Event']['OWC']
                        object_no = object_data[0]
                    else:
                        object_no = ""
                        owcloudurl = ""

                    # Calculate exposure using config values
                    mag_ref = config.get_mag_for_40ms_exposure()
                    extinction_mag = min(2, -0.5 + 0.5/math.cos((90-starAlt)*2*math.pi/360))
                    exposure = round(max(40, 40 * pow(2, round(combMag + extinction_mag - mag_ref + 0.5, 0)))/20)*20/1000.0

                    # Create dictionary of occultation events
                    occultation = {
                        'name': name + ' - ' + stationName, 
                        'station_name': stationName,
                        'ow_eventid': eventId, 
                        'id': name + ' : ' + star_id + ' : ' + stationName,
                        'ra': ra, 'dec': dec,
                        'star_mag': starMag, 'mag_drop': magDrop, 'comb_mag': combMag,
                        'event_time': eventCenterTime, 'start_time': startTime, 'end_time': endTime, 'goto_time': gotoTime,
                        'event_duration': eventDuration, 'event_uncertainty': eventUncertainty, 'recording_duration': recording_duration,
                        'occelmnt': eventOccelmnt, 'source': 'OWCloud',
                        'latitude': latitude, 'longitude': longitude, 'star_az': starAz, 'star_alt': starAlt,
                        'star_id': star_id, 'object_no': object_no, 'object_name': name, 'exposure': exposure
                    }
                    if owcloudurl:
                        occultation['owcloudurl'] = owcloudurl

                    occultations.append(occultation)

        return occultations

class TemplateManager:
    """Handles template file operations"""
    
    @staticmethod
    def find_template_files(template_folder=None):
        """Find all template files in the specified folder"""
        if template_folder is None:
            template_folder = config.get_file_folder()
            
        template_files = []
        try:
            if os.path.exists(template_folder):
                for file in os.listdir(template_folder):
                    if file.lower().endswith('.txt') and 'template' in file.lower():
                        template_files.append(file)
            template_files.sort()
        except Exception as e:
            print(f"Error finding template files: {e}")
        
        return template_files, template_folder
    
    @staticmethod
    def get_template_info(template_path):
        """Get template file information"""
        try:
            size = os.path.getsize(template_path)
            mtime = datetime.fromtimestamp(os.path.getmtime(template_path))
            return size, mtime
        except:
            return 0, datetime.min
    
    @staticmethod
    def load_template(template_path):
        """Load template content from file"""
        try:
            if template_path and os.path.exists(template_path):
                with open(template_path, 'r') as f:
                    return f.read()
            else:
                # Try default template
                default_path = config.get_full_file_path('SharpCap Owcloud template.txt')
                if os.path.exists(default_path):
                    with open(default_path, 'r') as f:
                        return f.read()
                else:
                    return None
        except Exception as e:
            print(f"Error loading template: {e}")
            return None

class OccultationEvent:
    """Represents a single occultation event with all calculations"""
    
    def __init__(self, event_data):
        self.original_data = event_data
        self.selected = True
        self.custom_exposure = None  # Track custom exposure settings
        self._parse_event_data(event_data)
        self._calculate_derived_values()
    
    def _parse_event_data(self, data):
        """Parse event data from OW Cloud JSON format"""
        self.name = data.get('name', '')
        self.station_name = data.get('station_name', '')
        self.ow_eventid = data.get('ow_eventid', '')
        self.event_id = data.get('id', '')
        self.object_name = data.get('object_name', '')
        
        self.ra = float(data.get('ra', 0))
        self.dec = float(data.get('dec', 0))
        self.ra_hours = self.ra
        self.dec_degrees = self.dec
        
        self.star_mag = float(data.get('star_mag', 0))
        self.mag_drop = float(data.get('mag_drop', 0))
        self.comb_mag = float(data.get('comb_mag', 0))
        self.magnitude = self.star_mag
        
        self.event_time = data.get('event_time', '')
        self.start_time_str = data.get('start_time', '')
        self.end_time_str = data.get('end_time', '')
        self.goto_time_str = data.get('goto_time', '')
        
        self.event_duration = float(data.get('event_duration', 0))
        self.event_uncertainty = float(data.get('event_uncertainty', 0))
        self.recording_duration = int(data.get('recording_duration', 0))
        
        self.star_id = data.get('star_id', '')
        self.star_az = float(data.get('star_az', 0))
        self.star_alt = float(data.get('star_alt', 0))
        
        self.object_no = data.get('object_no', '')
        self.asteroid_name = self.object_name
        self.star_name = self.star_id
        self.event_name = self.name
        
        self.latitude = float(data.get('latitude', 0))
        self.longitude = float(data.get('longitude', 0))
        self.precalc_exposure = float(data.get('exposure', 0))
        
        self.source = data.get('source', '')
        self.owcloudurl = data.get('owcloudurl', '')
        
        self.event_date = self._extract_date_from_iso(self.event_time)
        self.event_time_utc = self._extract_time_from_iso(self.event_time)
        
        self.duration_seconds = self.event_duration
        self.max_duration_seconds = self.event_duration
        self.uncertainty_seconds = self.event_uncertainty
    
    def _extract_date_from_iso(self, iso_string):
        """Extract date part from ISO datetime string"""
        if iso_string:
            return iso_string.split('T')[0]
        return ''
    
    def _extract_time_from_iso(self, iso_string):
        """Extract time part from ISO datetime string"""
        if iso_string:
            time_part = iso_string.split('T')[1] if 'T' in iso_string else iso_string
            return time_part.replace('Z', '')
        return ''
    
    def _calculate_derived_values(self):
        """Calculate exposure time and recording parameters"""
        if self.custom_exposure is not None:
            # Use custom exposure if set
            self.exposure_ms = int(self.custom_exposure * 1000)
        elif self.precalc_exposure > 0:
            self.exposure_ms = int(self.precalc_exposure * 1000)
        else:
            if self.star_mag > 0:
                mag_ref = config.get_mag_for_40ms_exposure()
                self.exposure_ms = max(10, int(40 * (2.5 ** (self.star_mag - mag_ref))))
                if self.exposure_ms > 1000:
                    self.exposure_ms = 1000
            else:
                self.exposure_ms = 40
        
        self.event_datetime = self._parse_iso_datetime(self.event_time)
        self.start_time = self._parse_iso_datetime(self.start_time_str)
        self.end_time = self._parse_iso_datetime(self.end_time_str)
        self.goto_time = self._parse_iso_datetime(self.goto_time_str)
        
        if self.recording_duration == 0 and self.start_time and self.end_time:
            duration_delta = self.end_time - self.start_time
            self.recording_duration = int(duration_delta.total_seconds())
    
    def set_custom_exposure(self, exposure_ms):
        """Set custom exposure in milliseconds"""
        self.custom_exposure = exposure_ms / 1000.0
        self.exposure_ms = exposure_ms
    
    def get_exposure_seconds(self):
        """Get exposure in seconds for template substitution"""
        return self.exposure_ms / 1000.0
    
    def has_custom_exposure(self):
        """Check if event has custom exposure setting"""
        return self.custom_exposure is not None
    
    def _parse_iso_datetime(self, iso_string):
        """Parse ISO format datetime string"""
        if not iso_string:
            return None
            
        try:
            formats = [
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%dT%H:%M:%S.%fZ",
                "%Y-%m-%d %H:%M:%S",
            ]
            
            clean_string = iso_string.replace('Z', '')
            
            for fmt in formats:
                try:
                    return datetime.strptime(clean_string, fmt)
                except ValueError:
                    continue
            
            print(f"Could not parse datetime: {iso_string}")
            return None
            
        except Exception as ex:
            print(f"Error parsing datetime '{iso_string}': {ex}")
            return None
    
    def get_coordinates_string(self):
        """Get formatted coordinate string"""
        return f"{self.ra:.4f}h, {self.dec:.4f}°"
    
    def get_status_info(self):
        """Get event status information"""
        if not self.event_datetime:
            return "Invalid Date"
        
        now = datetime.utcnow()
        if self.event_datetime < now:
            return "Past Event"
        
        time_to_event = self.event_datetime - now
        if time_to_event.total_seconds() < config.get_goto_lead_time():
            return "Starting Soon"
        
        days = time_to_event.days
        hours = time_to_event.seconds // 3600
        return f"{days}d {hours}h"

class OccultationManager:
    """Core manager class for GUI"""
    
    def __init__(self):
        self.events = []
        self.all_events = []
        self.selected_events = set()
        self.station_filter = ""
        self.running = False
        
        self.event_processor = EventProcessor()
        self.template_manager = TemplateManager()
    
    def load_events_from_files(self):
        """Load events from saved JSON files"""
        events_data = self.event_processor.load_occultations(config.get_latest_occultations_file())
        if not events_data:
            events_data = self.event_processor.load_occultations(config.get_occultations_file())
        
        if events_data:
            self.all_events = [OccultationEvent(event) for event in events_data]
            self.events = self.all_events[:]
            self.sort_events()
            return True
        return False
    
    def download_events_from_cloud(self):
        """Download events from OW Cloud"""
        try:
            events_data = self.event_processor.update_ow_cloud_events()
            if events_data:
                self.all_events = [OccultationEvent(event) for event in events_data]
                self.events = self.all_events[:]
                self.sort_events()
                return len(self.events)
            return 0
        except Exception as e:
            print(f"Error downloading events: {e}")
            return -1
    
    def sort_events(self):
        """Sort events by event time"""
        self.events.sort(key=lambda x: x.event_datetime if x.event_datetime else datetime.min)
    
    def get_filtered_events(self):
        """Get events filtered by station"""
        if self.station_filter:
            return [e for e in self.events if self.station_filter.lower() in e.station_name.lower()]
        return self.events
    
    def set_station_filter(self, filter_text):
        """Set station filter"""
        self.station_filter = filter_text
    
    def clear_station_filter(self):
        """Clear station filter"""
        self.station_filter = ""
        self.events = self.all_events[:]
    
    def select_all_events(self):
        """Select all filtered events"""
        filtered_events = self.get_filtered_events()
        self.selected_events.update(filtered_events)
        return len(filtered_events)
    
    def select_no_events(self):
        """Deselect all events"""
        self.selected_events.clear()
    
    def toggle_event_selection(self, event_indices):
        """Toggle selection of specific events by index"""
        filtered_events = self.get_filtered_events()
        for idx in event_indices:
            if 0 <= idx < len(filtered_events):
                event = filtered_events[idx]
                if event in self.selected_events:
                    self.selected_events.remove(event)
                else:
                    self.selected_events.add(event)
    
    def generate_sequences(self, template_path=None, progress_callback=None):
        """Generate sequence files for selected events"""
        selected_events = list(self.selected_events)
        if not selected_events:
            return 0, 0, "No events selected"
        
        template_content = self.template_manager.load_template(template_path)
        if not template_content:
            return 0, 0, "Template not found or empty"
        
        success_count = 0
        error_count = 0
        sequence_path = config.get_sequence_path()
        
        for i, event in enumerate(selected_events):
            try:
                if progress_callback:
                    progress_callback(i + 1, len(selected_events), f"Processing {event.event_name}")
                
                if save_occultation_sequence(event, template_path or "", sequence_path):
                    success_count += 1
                else:
                    error_count += 1
            except Exception as e:
                error_count += 1
                print(f"Error creating sequence for {event.event_name}: {e}")
        
        return success_count, error_count, f"Created {success_count} of {len(selected_events)} sequences"
    
    def generate_single_sequence(self, event, template_path=None):
        """Generate sequence for a single event"""
        template_content = self.template_manager.load_template(template_path)
        if not template_content:
            return False, "Template not found or empty"
        
        sequence_path = config.get_sequence_path()
        try:
            if save_occultation_sequence(event, template_path or "", sequence_path):
                return True, "Sequence generated successfully"
            else:
                return False, "Failed to save sequence file"
        except Exception as e:
            return False, f"Error creating sequence: {e}"

###############################
# GUI Dialog Classes
###############################

class ExposureEditDialog(Form):
    """Dialog for editing event exposure"""
    
    def __init__(self, event):
        Form.__init__(self)
        self.event = event
        self.new_exposure_ms = event.exposure_ms
        self.setup_ui()
    
    def setup_ui(self):
        """Setup exposure edit dialog UI"""
        self.Text = f"Edit Exposure - {self.event.event_name}"
        self.Size = Size(400, 300)
        self.StartPosition = FormStartPosition.CenterParent
        self.FormBorderStyle = FormBorderStyle.FixedDialog
        self.MaximizeBox = False
        self.MinimizeBox = False
        
        # Event info
        lbl_event = Label()
        lbl_event.Text = f"Event: {self.event.event_name}"
        lbl_event.Location = Point(20, 20)
        lbl_event.Size = Size(350, 20)
        lbl_event.Font = Font("Microsoft Sans Serif", 9, FontStyle.Bold)
        self.Controls.Add(lbl_event)
        
        lbl_star = Label()
        lbl_star.Text = f"Star: {self.event.star_name} (Mag: {self.event.star_mag:.1f})"
        lbl_star.Location = Point(20, 45)
        lbl_star.Size = Size(350, 20)
        self.Controls.Add(lbl_star)
        
        lbl_current = Label()
        current_text = f"Current Exposure: {self.event.exposure_ms} ms"
        if self.event.has_custom_exposure():
            current_text += " (Custom)"
        else:
            current_text += " (Calculated)"
        lbl_current.Text = current_text
        lbl_current.Location = Point(20, 70)
        lbl_current.Size = Size(350, 20)
        self.Controls.Add(lbl_current)
        
        # Exposure input
        lbl_new_exposure = Label()
        lbl_new_exposure.Text = "New Exposure (ms):"
        lbl_new_exposure.Location = Point(20, 110)
        lbl_new_exposure.Size = Size(120, 20)
        self.Controls.Add(lbl_new_exposure)
        
        self.txt_exposure = TextBox()
        self.txt_exposure.Text = str(self.event.exposure_ms)
        self.txt_exposure.Location = Point(150, 110)
        self.txt_exposure.Size = Size(100, 20)
        self.Controls.Add(self.txt_exposure)
        
        # Quick exposure buttons
        lbl_quick = Label()
        lbl_quick.Text = "Quick Settings:"
        lbl_quick.Location = Point(20, 150)
        lbl_quick.Size = Size(100, 20)
        self.Controls.Add(lbl_quick)
        
        quick_exposures = [10, 20, 40, 80, 100, 200, 500, 1000]
        x_pos = 20
        y_pos = 175
        
        for i, exp in enumerate(quick_exposures):
            btn = Button()
            btn.Text = f"{exp}ms"
            btn.Size = Size(45, 25)
            btn.Location = Point(x_pos, y_pos)
            btn.Tag = exp
            btn.Click += self.quick_exposure_click
            self.Controls.Add(btn)
            
            x_pos += 50
            if (i + 1) % 4 == 0:
                x_pos = 20
                y_pos += 30
        
        # Buttons
        btn_ok = Button()
        btn_ok.Text = "OK"
        btn_ok.DialogResult = DialogResult.OK
        btn_ok.Location = Point(220, 240)
        btn_ok.Size = Size(75, 25)
        btn_ok.Click += self.ok_click
        self.Controls.Add(btn_ok)
        
        btn_cancel = Button()
        btn_cancel.Text = "Cancel"
        btn_cancel.DialogResult = DialogResult.Cancel
        btn_cancel.Location = Point(305, 240)
        btn_cancel.Size = Size(75, 25)
        self.Controls.Add(btn_cancel)
        
        btn_reset = Button()
        btn_reset.Text = "Reset to Calculated"
        btn_reset.Location = Point(20, 240)
        btn_reset.Size = Size(120, 25)
        btn_reset.Click += self.reset_click
        self.Controls.Add(btn_reset)
        
        self.AcceptButton = btn_ok
        self.CancelButton = btn_cancel
    
    def quick_exposure_click(self, sender, e):
        """Handle quick exposure button click"""
        self.txt_exposure.Text = str(sender.Tag)
    
    def reset_click(self, sender, e):
        """Reset to calculated exposure"""
        # Temporarily clear custom exposure to get calculated value
        original_custom = self.event.custom_exposure
        self.event.custom_exposure = None
        self.event._calculate_derived_values()
        calculated_exposure = self.event.exposure_ms
        self.event.custom_exposure = original_custom
        self.event._calculate_derived_values()
        
        self.txt_exposure.Text = str(calculated_exposure)
    
    def ok_click(self, sender, e):
        """Handle OK button click"""
        try:
            self.new_exposure_ms = int(self.txt_exposure.Text)
            if self.new_exposure_ms < 1 or self.new_exposure_ms > 10000:
                MessageBox.Show("Exposure must be between 1 and 10000 ms", "Invalid Exposure", 
                              MessageBoxButtons.OK, MessageBoxIcon.Warning)
# Fixed bug here
                self.DialogResult = DialogResult.OK
                return
        except ValueError:
            MessageBox.Show("Please enter a valid number", "Invalid Input", 
                          MessageBoxButtons.OK, MessageBoxIcon.Warning)
            self.DialogResult = DialogResult.Ok
            return
    
    def get_new_exposure(self):
        """Get the new exposure value"""
        return self.new_exposure_ms

class ConfigurationDialog(Form):
    """Configuration dialog for GUI"""
    
    def __init__(self):
        Form.__init__(self)
        self.setup_ui()
        self.load_current_config()
    
    def setup_ui(self):
        """Setup configuration dialog UI"""
        self.Text = "Configuration Settings"
        self.Size = Size(600, 700)
        self.StartPosition = FormStartPosition.CenterParent
        self.FormBorderStyle = FormBorderStyle.FixedDialog
        self.MaximizeBox = False
        self.MinimizeBox = False
        
        # Create tabs
        tab_control = TabControl()
        tab_control.Location = Point(10, 10)
        tab_control.Size = Size(560, 600)
        self.Controls.Add(tab_control)
        
        # User Credentials Tab
        tab_credentials = TabPage()
        tab_credentials.Text = "Credentials"
        self.setup_credentials_tab(tab_credentials)
        tab_control.TabPages.Add(tab_credentials)
        
        # File Paths Tab
        tab_paths = TabPage()
        tab_paths.Text = "File Paths"
        self.setup_paths_tab(tab_paths)
        tab_control.TabPages.Add(tab_paths)
        
        # Recording Settings Tab
        tab_recording = TabPage()
        tab_recording.Text = "Recording"
        self.setup_recording_tab(tab_recording)
        tab_control.TabPages.Add(tab_recording)
        
        # API Settings Tab
        tab_api = TabPage()
        tab_api.Text = "API Settings"
        self.setup_api_tab(tab_api)
        tab_control.TabPages.Add(tab_api)
        
        # Observer/Telescope Tab
        tab_observer = TabPage()
        tab_observer.Text = "Observer/Telescope"
        self.setup_observer_tab(tab_observer)
        tab_control.TabPages.Add(tab_observer)
        
        # Buttons
        btn_ok = Button()
        btn_ok.Text = "Save"
        btn_ok.DialogResult = DialogResult.OK
        btn_ok.Location = Point(350, 630)
        btn_ok.Size = Size(75, 25)
        btn_ok.Click += self.save_config_click
        self.Controls.Add(btn_ok)
        
        btn_cancel = Button()
        btn_cancel.Text = "Cancel"
        btn_cancel.DialogResult = DialogResult.Cancel
        btn_cancel.Location = Point(435, 630)
        btn_cancel.Size = Size(75, 25)
        self.Controls.Add(btn_cancel)
        
        btn_reset = Button()
        btn_reset.Text = "Reset to Defaults"
        btn_reset.Location = Point(10, 630)
        btn_reset.Size = Size(120, 25)
        btn_reset.Click += self.reset_defaults_click
        self.Controls.Add(btn_reset)
        
        self.AcceptButton = btn_ok
        self.CancelButton = btn_cancel
    
    def setup_credentials_tab(self, tab):
        """Setup credentials tab"""
        lbl_email = Label()
        lbl_email.Text = "OWC Email:"
        lbl_email.Location = Point(20, 30)
        lbl_email.Size = Size(100, 20)
        tab.Controls.Add(lbl_email)
        
        self.txt_email = TextBox()
        self.txt_email.Location = Point(130, 30)
        self.txt_email.Size = Size(300, 20)
        tab.Controls.Add(self.txt_email)
        
        lbl_password = Label()
        lbl_password.Text = "OWC Password:"
        lbl_password.Location = Point(20, 60)
        lbl_password.Size = Size(100, 20)
        tab.Controls.Add(lbl_password)
        
        self.txt_password = TextBox()
        self.txt_password.Location = Point(130, 60)
        self.txt_password.Size = Size(300, 20)
        self.txt_password.UseSystemPasswordChar = True
        tab.Controls.Add(self.txt_password)
    
    def setup_paths_tab(self, tab):
        """Setup file paths tab"""
        lbl_file_folder = Label()
        lbl_file_folder.Text = "File Folder:"
        lbl_file_folder.Location = Point(20, 30)
        lbl_file_folder.Size = Size(100, 20)
        tab.Controls.Add(lbl_file_folder)
        
        self.txt_file_folder = TextBox()
        self.txt_file_folder.Location = Point(130, 30)
        self.txt_file_folder.Size = Size(250, 20)
        tab.Controls.Add(self.txt_file_folder)
        
        btn_browse_folder = Button()
        btn_browse_folder.Text = "Browse"
        btn_browse_folder.Location = Point(390, 29)
        btn_browse_folder.Size = Size(60, 22)
        btn_browse_folder.Click += self.browse_file_folder_click
        tab.Controls.Add(btn_browse_folder)
        
        lbl_sequence_path = Label()
        lbl_sequence_path.Text = "Sequence Path:"
        lbl_sequence_path.Location = Point(20, 60)
        lbl_sequence_path.Size = Size(100, 20)
        tab.Controls.Add(lbl_sequence_path)
        
        self.txt_sequence_path = TextBox()
        self.txt_sequence_path.Location = Point(130, 60)
        self.txt_sequence_path.Size = Size(250, 20)
        tab.Controls.Add(self.txt_sequence_path)
        
        btn_browse_sequence = Button()
        btn_browse_sequence.Text = "Browse"
        btn_browse_sequence.Location = Point(390, 59)
        btn_browse_sequence.Size = Size(60, 22)
        btn_browse_sequence.Click += self.browse_sequence_path_click
        tab.Controls.Add(btn_browse_sequence)
        
        lbl_occ_file = Label()
        lbl_occ_file.Text = "Occultations File:"
        lbl_occ_file.Location = Point(20, 90)
        lbl_occ_file.Size = Size(100, 20)
        tab.Controls.Add(lbl_occ_file)
        
        self.txt_occ_file = TextBox()
        self.txt_occ_file.Location = Point(130, 90)
        self.txt_occ_file.Size = Size(300, 20)
        tab.Controls.Add(self.txt_occ_file)
        
        lbl_latest_file = Label()
        lbl_latest_file.Text = "Latest File:"
        lbl_latest_file.Location = Point(20, 120)
        lbl_latest_file.Size = Size(100, 20)
        tab.Controls.Add(lbl_latest_file)
        
        self.txt_latest_file = TextBox()
        self.txt_latest_file.Location = Point(130, 120)
        self.txt_latest_file.Size = Size(300, 20)
        tab.Controls.Add(self.txt_latest_file)
    
    def setup_recording_tab(self, tab):
        """Setup recording settings tab"""
        lbl_base_duration = Label()
        lbl_base_duration.Text = "Base Duration (s):"
        lbl_base_duration.Location = Point(20, 30)
        lbl_base_duration.Size = Size(120, 20)
        tab.Controls.Add(lbl_base_duration)
        
        self.txt_base_duration = TextBox()
        self.txt_base_duration.Location = Point(150, 30)
        self.txt_base_duration.Size = Size(100, 20)
        tab.Controls.Add(self.txt_base_duration)
        
        lbl_goto_lead = Label()
        lbl_goto_lead.Text = "GOTO Lead Time (s):"
        lbl_goto_lead.Location = Point(20, 60)
        lbl_goto_lead.Size = Size(120, 20)
        tab.Controls.Add(lbl_goto_lead)
        
        self.txt_goto_lead = TextBox()
        self.txt_goto_lead.Location = Point(150, 60)
        self.txt_goto_lead.Size = Size(100, 20)
        tab.Controls.Add(self.txt_goto_lead)
        
        lbl_mag_exposure = Label()
        lbl_mag_exposure.Text = "Mag for 40ms exp:"
        lbl_mag_exposure.Location = Point(20, 90)
        lbl_mag_exposure.Size = Size(120, 20)
        tab.Controls.Add(lbl_mag_exposure)
        
        self.txt_mag_exposure = TextBox()
        self.txt_mag_exposure.Location = Point(150, 90)
        self.txt_mag_exposure.Size = Size(100, 20)
        tab.Controls.Add(self.txt_mag_exposure)
    
    def setup_api_tab(self, tab):
        """Setup API settings tab"""
        lbl_host = Label()
        lbl_host.Text = "API Host:"
        lbl_host.Location = Point(20, 30)
        lbl_host.Size = Size(100, 20)
        tab.Controls.Add(lbl_host)
        
        self.txt_host = TextBox()
        self.txt_host.Location = Point(130, 30)
        self.txt_host.Size = Size(300, 20)
        tab.Controls.Add(self.txt_host)
        
        lbl_api_key = Label()
        lbl_api_key.Text = "API Key:"
        lbl_api_key.Location = Point(20, 60)
        lbl_api_key.Size = Size(100, 20)
        tab.Controls.Add(lbl_api_key)
        
        self.txt_api_key = TextBox()
        self.txt_api_key.Location = Point(130, 60)
        self.txt_api_key.Size = Size(300, 20)
        tab.Controls.Add(self.txt_api_key)
    
    def setup_observer_tab(self, tab):
        """Setup observer/telescope tab"""
        # Observer section
        lbl_observer_section = Label()
        lbl_observer_section.Text = "Observer Information (for NA Report Form):"
        lbl_observer_section.Location = Point(20, 10)
        lbl_observer_section.Size = Size(400, 20)
        lbl_observer_section.Font = Font(lbl_observer_section.Font, FontStyle.Bold)
        tab.Controls.Add(lbl_observer_section)
        
        # Observer Name
        lbl_observer_name = Label()
        lbl_observer_name.Text = "Observer Name:"
        lbl_observer_name.Location = Point(20, 40)
        lbl_observer_name.Size = Size(120, 20)
        tab.Controls.Add(lbl_observer_name)
        
        self.txt_observer_name = TextBox()
        self.txt_observer_name.Location = Point(150, 40)
        self.txt_observer_name.Size = Size(300, 20)
        tab.Controls.Add(self.txt_observer_name)
        
        # Observer Email
        lbl_observer_email = Label()
        lbl_observer_email.Text = "Observer Email:"
        lbl_observer_email.Location = Point(20, 70)
        lbl_observer_email.Size = Size(120, 20)
        tab.Controls.Add(lbl_observer_email)
        
        self.txt_observer_email = TextBox()
        self.txt_observer_email.Location = Point(150, 70)
        self.txt_observer_email.Size = Size(300, 20)
        tab.Controls.Add(self.txt_observer_email)
        
        # Observer Latitude
        lbl_latitude = Label()
        lbl_latitude.Text = "Latitude (deg):"
        lbl_latitude.Location = Point(20, 100)
        lbl_latitude.Size = Size(120, 20)
        tab.Controls.Add(lbl_latitude)
        
        self.txt_latitude = TextBox()
        self.txt_latitude.Location = Point(150, 100)
        self.txt_latitude.Size = Size(150, 20)
        tab.Controls.Add(self.txt_latitude)
        
        # Observer Longitude
        lbl_longitude = Label()
        lbl_longitude.Text = "Longitude (deg):"
        lbl_longitude.Location = Point(20, 130)
        lbl_longitude.Size = Size(120, 20)
        tab.Controls.Add(lbl_longitude)
        
        self.txt_longitude = TextBox()
        self.txt_longitude.Location = Point(150, 130)
        self.txt_longitude.Size = Size(150, 20)
        tab.Controls.Add(self.txt_longitude)
        
        # Observer Elevation
        lbl_elevation = Label()
        lbl_elevation.Text = "Elevation (m):"
        lbl_elevation.Location = Point(20, 160)
        lbl_elevation.Size = Size(120, 20)
        tab.Controls.Add(lbl_elevation)
        
        self.txt_elevation = TextBox()
        self.txt_elevation.Location = Point(150, 160)
        self.txt_elevation.Size = Size(150, 20)
        tab.Controls.Add(self.txt_elevation)
        
        # Telescope section
        lbl_telescope_section = Label()
        lbl_telescope_section.Text = "Telescope Information (for NA Report Form):"
        lbl_telescope_section.Location = Point(20, 200)
        lbl_telescope_section.Size = Size(400, 20)
        lbl_telescope_section.Font = Font(lbl_telescope_section.Font, FontStyle.Bold)
        tab.Controls.Add(lbl_telescope_section)
        
        # Telescope Aperture
        lbl_aperture = Label()
        lbl_aperture.Text = "Aperture (mm):"
        lbl_aperture.Location = Point(20, 230)
        lbl_aperture.Size = Size(120, 20)
        tab.Controls.Add(lbl_aperture)
        
        self.txt_aperture = TextBox()
        self.txt_aperture.Location = Point(150, 230)
        self.txt_aperture.Size = Size(150, 20)
        tab.Controls.Add(self.txt_aperture)
        
        # Telescope Focal Length
        lbl_focal_length = Label()
        lbl_focal_length.Text = "Focal Length (mm):"
        lbl_focal_length.Location = Point(20, 260)
        lbl_focal_length.Size = Size(120, 20)
        tab.Controls.Add(lbl_focal_length)
        
        self.txt_focal_length = TextBox()
        self.txt_focal_length.Location = Point(150, 260)
        self.txt_focal_length.Size = Size(150, 20)
        tab.Controls.Add(self.txt_focal_length)
        
        # Telescope Type
        lbl_telescope_type = Label()
        lbl_telescope_type.Text = "Telescope Type:"
        lbl_telescope_type.Location = Point(20, 290)
        lbl_telescope_type.Size = Size(120, 20)
        tab.Controls.Add(lbl_telescope_type)
        
        self.cmb_telescope_type = ComboBox()
        self.cmb_telescope_type.Location = Point(150, 290)
        self.cmb_telescope_type.Size = Size(300, 20)
        self.cmb_telescope_type.DropDownStyle = ComboBoxStyle.DropDownList
        self.cmb_telescope_type.Items.AddRange(['SCT including Cass and Mak', 'Newtonian', 'Refractor', 'Dobsonian'])
        tab.Controls.Add(self.cmb_telescope_type)
    
    def load_current_config(self):
        """Load current configuration into controls"""
        self.txt_email.Text = config.get_owc_email()
        self.txt_password.Text = config.get_owc_password()
        self.txt_file_folder.Text = config.get_file_folder()
        self.txt_sequence_path.Text = config.get_sequence_path()
        self.txt_occ_file.Text = config.get_occultations_file()
        self.txt_latest_file.Text = config.get_latest_occultations_file()
        self.txt_base_duration.Text = str(config.get_base_duration())
        self.txt_goto_lead.Text = str(config.get_goto_lead_time())
        self.txt_mag_exposure.Text = str(config.get_mag_for_40ms_exposure())
        self.txt_host.Text = config.get_host()
        self.txt_api_key.Text = config.get_api_key()
        
        # Observer/Telescope settings
        self.txt_observer_name.Text = config.get_observer_name()
        self.txt_observer_email.Text = config.get_observer_email()
        self.txt_latitude.Text = str(config.get_observer_latitude())
        self.txt_longitude.Text = str(config.get_observer_longitude())
        self.txt_elevation.Text = str(config.get_observer_elevation())
        self.txt_aperture.Text = str(config.get_telescope_aperture())
        self.txt_focal_length.Text = str(config.get_telescope_focal_length())
        self.cmb_telescope_type.Text = config.get_telescope_type()
    
    def browse_file_folder_click(self, sender, e):
        """Browse for file folder"""
        dialog = FolderBrowserDialog()
        dialog.SelectedPath = self.txt_file_folder.Text
        if dialog.ShowDialog() == DialogResult.OK:
            self.txt_file_folder.Text = dialog.SelectedPath
    
    def browse_sequence_path_click(self, sender, e):
        """Browse for sequence path"""
        dialog = FolderBrowserDialog()
        dialog.SelectedPath = self.txt_sequence_path.Text
        if dialog.ShowDialog() == DialogResult.OK:
            self.txt_sequence_path.Text = dialog.SelectedPath
    
    def save_config_click(self, sender, e):
        """Save configuration"""
        try:
            # Update config with form values
            config.set_owc_email(self.txt_email.Text)
            config.set_owc_password(self.txt_password.Text)
            config.set_file_folder(self.txt_file_folder.Text)
            config.set_sequence_path(self.txt_sequence_path.Text)
            config.set_occultations_file(self.txt_occ_file.Text)
            config.set_latest_occultations_file(self.txt_latest_file.Text)
            config.set_base_duration(int(self.txt_base_duration.Text))
            config.set_goto_lead_time(int(self.txt_goto_lead.Text))
            config.set_mag_for_40ms_exposure(float(self.txt_mag_exposure.Text))
            config.set_host(self.txt_host.Text)
            config.set_api_key(self.txt_api_key.Text)
            
            # Observer/Telescope settings
            config.set_observer_name(self.txt_observer_name.Text)
            config.set_observer_email(self.txt_observer_email.Text)
            config.set_observer_latitude(float(self.txt_latitude.Text) if self.txt_latitude.Text else 0.0)
            config.set_observer_longitude(float(self.txt_longitude.Text) if self.txt_longitude.Text else 0.0)
            config.set_observer_elevation(float(self.txt_elevation.Text) if self.txt_elevation.Text else 0.0)
            config.set_telescope_aperture(int(self.txt_aperture.Text) if self.txt_aperture.Text else 0)
            config.set_telescope_focal_length(int(self.txt_focal_length.Text) if self.txt_focal_length.Text else 0)
            config.set_telescope_type(self.cmb_telescope_type.Text)
            
            # Validate and save
            errors = config.validate_config()
            if errors:
                MessageBox.Show("Configuration errors:\n" + "\n".join(errors), 
                              "Configuration Error", MessageBoxButtons.OK, MessageBoxIcon.Warning)
                return
            
            if config.save_config():
                MessageBox.Show("Configuration saved successfully!", "Success", 
                              MessageBoxButtons.OK, MessageBoxIcon.Information)
            else:
                MessageBox.Show("Failed to save configuration!", "Error", 
                              MessageBoxButtons.OK, MessageBoxIcon.Error)
                
        except ValueError as e:
            MessageBox.Show(f"Invalid numeric value: {e}", "Input Error", 
                          MessageBoxButtons.OK, MessageBoxIcon.Error)
        except Exception as e:
            MessageBox.Show(f"Error saving configuration: {e}", "Error", 
                          MessageBoxButtons.OK, MessageBoxIcon.Error)
    
    def reset_defaults_click(self, sender, e):
        """Reset to default configuration"""
        if MessageBox.Show("Reset all settings to defaults?", "Confirm Reset", 
                         MessageBoxButtons.YesNo, MessageBoxIcon.Question) == DialogResult.Yes:
            config.reset_to_defaults()
            self.load_current_config()
            MessageBox.Show("Configuration reset to defaults", "Reset Complete", 
                          MessageBoxButtons.OK, MessageBoxIcon.Information)

class TemplateSelectionDialog(Form):
    """Dialog for selecting sequence template"""
    
    def __init__(self):
        Form.__init__(self)
        self.selected_template_path = ""
        self.template_manager = TemplateManager()
        self.setup_ui()
    
    def setup_ui(self):
        """Setup template selection UI"""
        self.Text = "Select Sequence Template"
        self.Size = Size(600, 500)
        self.StartPosition = FormStartPosition.CenterParent
        self.FormBorderStyle = FormBorderStyle.FixedDialog
        self.MaximizeBox = False
        self.MinimizeBox = False
        
        # Template list
        lbl_templates = Label()
        lbl_templates.Text = "Available Templates:"
        lbl_templates.Location = Point(10, 10)
        lbl_templates.Size = Size(200, 20)
        self.Controls.Add(lbl_templates)
        
        self.lst_templates = ListBox()
        self.lst_templates.Location = Point(10, 35)
        self.lst_templates.Size = Size(560, 200)
        self.lst_templates.SelectionMode = SelectionMode.One
        self.Controls.Add(self.lst_templates)
        
        # Load templates
        self.load_templates()
        
        # Template preview
        lbl_preview = Label()
        lbl_preview.Text = "Template Preview:"
        lbl_preview.Location = Point(10, 250)
        lbl_preview.Size = Size(200, 20)
        self.Controls.Add(lbl_preview)
        
        self.txt_preview = TextBox()
        self.txt_preview.Multiline = True
        self.txt_preview.ReadOnly = True
        self.txt_preview.ScrollBars = ScrollBars.Vertical
        self.txt_preview.Location = Point(10, 275)
        self.txt_preview.Size = Size(560, 150)
        self.Controls.Add(self.txt_preview)
        
        # Buttons
        btn_ok = Button()
        btn_ok.Text = "OK"
        btn_ok.DialogResult = DialogResult.OK
        btn_ok.Location = Point(430, 440)
        btn_ok.Size = Size(75, 25)
        self.Controls.Add(btn_ok)
        
        btn_cancel = Button()
        btn_cancel.Text = "Cancel"
        btn_cancel.DialogResult = DialogResult.Cancel
        btn_cancel.Location = Point(515, 440)
        btn_cancel.Size = Size(75, 25)
        self.Controls.Add(btn_cancel)
        
        # Wire events
        self.lst_templates.SelectedIndexChanged += self.template_selected
        
        self.AcceptButton = btn_ok
        self.CancelButton = btn_cancel
    
    def load_templates(self):
        """Load available templates into the list"""
        template_files, template_folder = self.template_manager.find_template_files()
        
        # Add default option
        #self.lst_templates.Items.Add("Default Template")
        
        # Add template files
        for template_file in template_files:
            template_path = os.path.join(template_folder, template_file)
            size, mtime = self.template_manager.get_template_info(template_path)
            display_text = f"{template_file} ({size} bytes, {mtime.strftime('%Y-%m-%d %H:%M')})"
            self.lst_templates.Items.Add(display_text)
        
        # Select first item
        if self.lst_templates.Items.Count > 0:
            self.lst_templates.SelectedIndex = 0
    
    def template_selected(self, sender, e):
        """Handle template selection change"""
        if self.lst_templates.SelectedIndex >= 0:
            if self.lst_templates.SelectedIndex == 0:
                # Default template
                self.selected_template_path = ""
                template_content = self.template_manager.load_template("")
            else:
                # Specific template file
                template_files, template_folder = self.template_manager.find_template_files()
                if self.lst_templates.SelectedIndex - 1 < len(template_files):
                    template_file = template_files[self.lst_templates.SelectedIndex - 1]
                    self.selected_template_path = os.path.join(template_folder, template_file)
                    template_content = self.template_manager.load_template(self.selected_template_path)
            
            # Show preview
            if template_content:
                # Show first 1000 characters
                preview = template_content[:1000]
                if len(template_content) > 1000:
                    preview += "\n\n... (truncated)"
                self.txt_preview.Text = preview
            else:
                self.txt_preview.Text = "Could not load template content"
    
    def get_selected_template_path(self):
        """Get the selected template path"""
        return self.selected_template_path

class EventsDataGrid(DataGridView):
    """Custom DataGridView for displaying occultation events"""
    
    def __init__(self):
        DataGridView.__init__(self)
        self.events = []
        self.setup_grid()
    
    def setup_grid(self):
        """Setup the data grid columns and properties"""
        self.AutoGenerateColumns = False
        self.AllowUserToAddRows = False
        self.AllowUserToDeleteRows = False
        self.SelectionMode = DataGridViewSelectionMode.FullRowSelect
        self.MultiSelect = True
        
        columns = [
            ("Selected", "Selected", 60, True),
            ("Event Name", "EventName", 150, False),
            ("Star", "StarName", 100, False),
            ("Asteroid", "AsteroidName", 120, False),
            ("Date/Time UTC", "DateTime", 120, False),
            ("Magnitude", "Magnitude", 80, False),
            ("Exposure (ms)", "ExposureMs", 90, False),
            ("Duration (s)", "Duration", 80, False),
            ("Coordinates", "Coordinates", 120, False),
            ("Status", "Status", 100, False)
        ]
        
        for name, data_name, width, editable in columns:
            if name == "Selected":
                col = DataGridViewCheckBoxColumn()
            else:
                col = DataGridViewTextBoxColumn()
            
            col.Name = data_name
            col.HeaderText = name
            col.Width = width
            col.ReadOnly = not editable
            self.Columns.Add(col)
        
        # Handle cell double-click for exposure editing
        self.CellDoubleClick += self.cell_double_click
    
    def cell_double_click(self, sender, e):
        """Handle cell double click for exposure editing"""
        if e.RowIndex >= 0 and e.ColumnIndex >= 0:
            if self.Columns[e.ColumnIndex].Name == "ExposureMs":
                event = self.Rows[e.RowIndex].Tag
                if event:
                    # Get parent form to handle exposure editing
                    parent_form = self.FindForm()
                    if hasattr(parent_form, 'edit_event_exposure'):
                        parent_form.edit_event_exposure(event)
    
    def update_events(self, events):
        """Update the grid with new events data"""
        self.events = events
        self.Rows.Clear()
        
        for event in events:
            row = self.Rows[self.Rows.Add()]
            row.Cells["Selected"].Value = event.selected
            row.Cells["EventName"].Value = event.event_name
            row.Cells["StarName"].Value = event.star_name
            row.Cells["AsteroidName"].Value = event.asteroid_name
            row.Cells["DateTime"].Value = f"{event.event_date} {event.event_time_utc}" if event.event_date else "N/A"
            row.Cells["Magnitude"].Value = f"{event.magnitude:.1f}" if event.magnitude > 0 else "N/A"
            
            # Show custom exposure indicator
            exposure_text = str(event.exposure_ms)
            if event.has_custom_exposure():
                exposure_text += "*"
            row.Cells["ExposureMs"].Value = exposure_text
            
            row.Cells["Duration"].Value = str(event.recording_duration)
            row.Cells["Coordinates"].Value = event.get_coordinates_string()
            row.Cells["Status"].Value = event.get_status_info()
            row.Tag = event
    
    def get_selected_events(self):
        """Get list of selected events"""
        selected = []
        for row in self.Rows:
            if row.Cells["Selected"].Value:
                selected.append(row.Tag)
        return selected
    
    def select_all_events(self, select=True):
        """Select or deselect all events"""
        for row in self.Rows:
            row.Cells["Selected"].Value = select
            if row.Tag:
                row.Tag.selected = select

class OccultationManagerGUI(Form):
    """Main GUI window for occultation management"""
    
    def __init__(self):
        Form.__init__(self)
        self.manager = OccultationManager()
        self.setup_ui()
        self.load_initial_data()
    
    def setup_ui(self):
        """Setup the main user interface"""
        self.Text = "Occultation Manager - SharpCap Integration"
        self.Size = Size(1200, 700)
        self.StartPosition = FormStartPosition.CenterScreen
        
        # Create menu bar
        menu_bar = self.create_menu_bar()
        self.MainMenuStrip = menu_bar
        self.Controls.Add(menu_bar)
        
        main_panel = Panel()
        main_panel.Dock = DockStyle.Fill
        self.Controls.Add(main_panel)
        
        toolbar = self.create_toolbar()
        toolbar.Parent = main_panel
        
        self.events_grid = EventsDataGrid()
        self.events_grid.Location = Point(10, 135)
        self.events_grid.Size = Size(1160, 400)
        self.events_grid.Anchor = AnchorStyles.Top | AnchorStyles.Left | AnchorStyles.Right | AnchorStyles.Bottom
        main_panel.Controls.Add(self.events_grid)
        
        bottom_panel = self.create_bottom_panel()
        bottom_panel.Parent = main_panel
        
        status_bar = self.create_status_bar()
        status_bar.Parent = main_panel
    
    def create_menu_bar(self):
        """Create the menu bar"""
        menu_bar = MenuStrip()
        
        # File menu
        menu_file = ToolStripMenuItem("File")
        menu_file.DropDownItems.Add(ToolStripMenuItem("Download Events", None, self.download_events_click))
        menu_file.DropDownItems.Add(ToolStripMenuItem("Refresh Events", None, self.refresh_events_click))
        menu_file.DropDownItems.Add(ToolStripSeparator())
        menu_file.DropDownItems.Add(ToolStripMenuItem("Exit", None, self.exit_click))
        menu_bar.Items.Add(menu_file)
        
        # Tools menu
        menu_tools = ToolStripMenuItem("Tools")
        menu_tools.DropDownItems.Add(ToolStripMenuItem("Configuration", None, self.show_configuration_click))
        menu_tools.DropDownItems.Add(ToolStripMenuItem("Template Manager", None, self.show_template_manager_click))
        menu_bar.Items.Add(menu_tools)
        
        # Help menu
        menu_help = ToolStripMenuItem("Help")
        menu_help.DropDownItems.Add(ToolStripMenuItem("About", None, self.show_about_click))
        menu_bar.Items.Add(menu_help)
        
        return menu_bar
    
    def create_toolbar(self):
        """Create the main toolbar"""
        toolbar = Panel()
        toolbar.Height = 50
        toolbar.Dock = DockStyle.Top
        toolbar.BackColor = SystemColors.Control
        
        btn_download = Button()
        btn_download.Text = "Download Events"
        btn_download.Size = Size(120, 25)
        btn_download.Location = Point(5, 25)
        btn_download.Click += self.download_events_click
        toolbar.Controls.Add(btn_download)
        
        btn_refresh = Button()
        btn_refresh.Text = "Refresh"
        btn_refresh.Size = Size(80, 25)
        btn_refresh.Location = Point(130, 25)
        btn_refresh.Click += self.refresh_events_click
        toolbar.Controls.Add(btn_refresh)
        
        btn_select_all = Button()
        btn_select_all.Text = "Select All"
        btn_select_all.Size = Size(80, 25)
        btn_select_all.Location = Point(220, 25)
        btn_select_all.Click += self.select_all_click
        toolbar.Controls.Add(btn_select_all)
        
        btn_select_none = Button()
        btn_select_none.Text = "Select None"
        btn_select_none.Size = Size(80, 25)
        btn_select_none.Location = Point(305, 25)
        btn_select_none.Click += self.select_none_click
        toolbar.Controls.Add(btn_select_none)
        
        btn_create_sequences = Button()
        btn_create_sequences.Text = "Create Sequences"
        btn_create_sequences.Size = Size(120, 25)
        btn_create_sequences.Location = Point(400, 25)
        btn_create_sequences.Click += self.create_sequences_click
        toolbar.Controls.Add(btn_create_sequences)
        
        btn_edit_exposure = Button()
        btn_edit_exposure.Text = "Edit Exposure"
        btn_edit_exposure.Size = Size(100, 25)
        btn_edit_exposure.Location = Point(530, 25)
        btn_edit_exposure.Click += self.edit_exposure_click
        toolbar.Controls.Add(btn_edit_exposure)
        
        btn_generate_report = Button()
        btn_generate_report.Text = "Generate Report"
        btn_generate_report.Size = Size(120, 25)
        btn_generate_report.Location = Point(640, 25)
        btn_generate_report.Click += self.generate_report_click
        toolbar.Controls.Add(btn_generate_report)
        
        return toolbar
    
    def create_bottom_panel(self):
        """Create the bottom control panel"""
        panel = Panel()
        panel.Height = 120
        panel.Dock = DockStyle.Bottom
        panel.BackColor = SystemColors.Control
        
        details_group = GroupBox()
        details_group.Text = "Event Details"
        details_group.Location = Point(10, 5)
        details_group.Size = Size(400, 110)
        panel.Controls.Add(details_group)
        
        self.lbl_event_details = Label()
        self.lbl_event_details.Text = "Select an event to view details"
        self.lbl_event_details.Location = Point(10, 20)
        self.lbl_event_details.Size = Size(380, 80)
        details_group.Controls.Add(self.lbl_event_details)
        
        config_group = GroupBox()
        config_group.Text = "Configuration"
        config_group.Location = Point(420, 5)
        config_group.Size = Size(300, 110)
        panel.Controls.Add(config_group)
        
        lbl_seq_path = Label()
        lbl_seq_path.Text = "Sequence Path:"
        lbl_seq_path.Location = Point(10, 25)
        lbl_seq_path.Size = Size(80, 20)
        config_group.Controls.Add(lbl_seq_path)
        
        self.txt_sequence_path = TextBox()
        self.txt_sequence_path.Text = config.get_sequence_path()
        self.txt_sequence_path.Location = Point(10, 45)
        self.txt_sequence_path.Size = Size(200, 20)
        config_group.Controls.Add(self.txt_sequence_path)
        
        btn_browse = Button()
        btn_browse.Text = "Browse"
        btn_browse.Location = Point(220, 44)
        btn_browse.Size = Size(60, 22)
        btn_browse.Click += self.browse_sequence_path_click
        config_group.Controls.Add(btn_browse)
        
        actions_group = GroupBox()
        actions_group.Text = "Filters"
        actions_group.Location = Point(730, 5)
        actions_group.Size = Size(200, 110)
        panel.Controls.Add(actions_group)
        
        btn_filter_today = Button()
        btn_filter_today.Text = "Today's Events"
        btn_filter_today.Location = Point(10, 25)
        btn_filter_today.Size = Size(130, 25)
        btn_filter_today.Click += self.filter_today_click
        actions_group.Controls.Add(btn_filter_today)
        
        btn_filter_upcoming = Button()
        btn_filter_upcoming.Text = "Upcoming"
        btn_filter_upcoming.Location = Point(10, 55)
        btn_filter_upcoming.Size = Size(130, 25)
        btn_filter_upcoming.Click += self.filter_upcoming_click
        actions_group.Controls.Add(btn_filter_upcoming)
        
        btn_show_all = Button()
        btn_show_all.Text = "Show All"
        btn_show_all.Location = Point(10, 85)
        btn_show_all.Size = Size(130, 25)
        btn_show_all.Click += self.show_all_click
        actions_group.Controls.Add(btn_show_all)
        
        self.events_grid.SelectionChanged += self.grid_selection_changed
        
        return panel
    
    def create_status_bar(self):
        """Create the status bar"""
        status_bar = Panel()
        status_bar.Height = 25
        status_bar.Dock = DockStyle.Bottom
        status_bar.BackColor = SystemColors.ControlDark
        
        self.lbl_status = Label()
        self.lbl_status.Text = "Ready"
        self.lbl_status.Location = Point(10, 5)
        self.lbl_status.Size = Size(400, 15)
        status_bar.Controls.Add(self.lbl_status)
        
        self.lbl_event_count = Label()
        self.lbl_event_count.Text = "0 events"
        self.lbl_event_count.Location = Point(500, 5)
        self.lbl_event_count.Size = Size(100, 15)
        status_bar.Controls.Add(self.lbl_event_count)
        
        return status_bar
    
    def load_initial_data(self):
        """Load initial events data"""
        self.update_status("Loading events...")
        if self.manager.load_events_from_files():
            self.refresh_display()
            self.update_status("Events loaded successfully")
        else:
            self.update_status("No events found - use Download Events to fetch from OW Cloud")
    
    def refresh_display(self):
        """Refresh the events display"""
        self.events_grid.update_events(self.manager.get_filtered_events())
        self.lbl_event_count.Text = f"{len(self.manager.get_filtered_events())} events"
    
    def update_status(self, message):
        """Update the status bar"""
        self.lbl_status.Text = message
        Application.DoEvents()
    
    # Event Handlers
    def download_events_click(self, sender, e):
        """Handle download events button click"""
        self.update_status("Downloading events from OW Cloud...")
        try:
            result = self.manager.download_events_from_cloud()
            if result > 0:
                self.refresh_display()
                self.update_status(f"Downloaded {result} events")
            elif result == 0:
                self.update_status("No events downloaded")
            else:
                self.update_status("Error downloading events")
        except Exception as ex:
            self.update_status(f"Error downloading events: {ex}")
            MessageBox.Show(f"Error downloading events: {ex}", "Download Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
    
    def refresh_events_click(self, sender, e):
        """Handle refresh button click"""
        self.load_initial_data()
    
    def select_all_click(self, sender, e):
        """Handle select all button click"""
        self.events_grid.select_all_events(True)
        for event in self.manager.get_filtered_events():
            self.manager.selected_events.add(event)
    
    def select_none_click(self, sender, e):
        """Handle select none button click"""
        self.events_grid.select_all_events(False)
        self.manager.selected_events.clear()
    
    def edit_exposure_click(self, sender, e):
        """Handle edit exposure button click"""
        selected_rows = []
        for row in self.events_grid.SelectedRows:
            selected_rows.append(row)
        
        if len(selected_rows) == 0:
            MessageBox.Show("Please select an event to edit exposure.", "No Event Selected", 
                          MessageBoxButtons.OK, MessageBoxIcon.Information)
            return
        elif len(selected_rows) > 1:
            MessageBox.Show("Please select only one event to edit exposure.", "Multiple Events Selected", 
                          MessageBoxButtons.OK, MessageBoxIcon.Information)
            return
        
        event = selected_rows[0].Tag
        if event:
            self.edit_event_exposure(event)
    
    def edit_event_exposure(self, event):
        """Edit exposure for a specific event"""
        exposure_dialog = ExposureEditDialog(event)
        if exposure_dialog.ShowDialog() == DialogResult.OK:
            new_exposure = exposure_dialog.get_new_exposure()
            event.set_custom_exposure(new_exposure)
            
            # Refresh the grid to show updated exposure
            self.refresh_display()
            
            # Ask if user wants to regenerate sequence
            result = MessageBox.Show(
                f"Exposure updated to {new_exposure}ms.\n\nWould you like to regenerate the sequence file for this event?",
                "Regenerate Sequence?",
                MessageBoxButtons.YesNo,
                MessageBoxIcon.Question
            )
            
            if result == DialogResult.Yes:
                self.regenerate_single_sequence(event)
    
    def regenerate_single_sequence(self, event):
        """Regenerate sequence for a single event"""
        template_dialog = TemplateSelectionDialog()
        if template_dialog.ShowDialog() == DialogResult.OK:
            template_path = template_dialog.get_selected_template_path()
            
            self.update_status(f"Generating sequence for {event.event_name}...")
            success, message = self.manager.generate_single_sequence(event, template_path)
            
            if success:
                self.update_status("Sequence generated successfully")
                MessageBox.Show(f"Sequence file regenerated successfully for {event.event_name}", 
                              "Success", MessageBoxButtons.OK, MessageBoxIcon.Information)
            else:
                self.update_status(f"Error: {message}")
                MessageBox.Show(f"Failed to regenerate sequence: {message}", 
                              "Error", MessageBoxButtons.OK, MessageBoxIcon.Error)
    
    def create_sequences_click(self, sender, e):
        """Handle create sequences button click"""
        selected_events = self.events_grid.get_selected_events()
        if not selected_events:
            MessageBox.Show("Please select events to create sequences for.", "No Events Selected", 
                          MessageBoxButtons.OK, MessageBoxIcon.Information)
            return
        
        self.manager.selected_events = set(selected_events)
        
        template_dialog = TemplateSelectionDialog()
        if template_dialog.ShowDialog() == DialogResult.OK:
            template_path = template_dialog.get_selected_template_path()
            self.create_sequences_for_events(template_path)
    
    def create_sequences_for_events(self, template_path):
        """Create sequence files for selected events"""
        sequence_path = self.txt_sequence_path.Text
        config.set_sequence_path(sequence_path)
        
        def progress_callback(current, total, message):
            self.update_status(f"Creating sequences... {current}/{total}")
        
        success_count, error_count, message = self.manager.generate_sequences(template_path, progress_callback)
        
        self.update_status(message)
        MessageBox.Show(f"Successfully created {success_count} of {success_count + error_count} sequence files.", 
                       "Sequence Creation Complete", MessageBoxButtons.OK, MessageBoxIcon.Information)
    
    def filter_today_click(self, sender, e):
        """Filter events for today"""
        today = datetime.utcnow().date()
        filtered_events = []
        for event in self.manager.all_events:
            if event.event_datetime and event.event_datetime.date() == today:
                filtered_events.append(event)
        
        self.manager.events = filtered_events
        self.refresh_display()
        self.update_status(f"Showing today's events: {len(filtered_events)}")
    
    def filter_upcoming_click(self, sender, e):
        """Filter upcoming events"""
        now = datetime.utcnow()
        filtered_events = []
        for event in self.manager.all_events:
            if event.event_datetime and event.event_datetime > now:
                filtered_events.append(event)
        
        self.manager.events = filtered_events
        self.refresh_display()
        self.update_status(f"Showing upcoming events: {len(filtered_events)}")
    
    def show_all_click(self, sender, e):
        """Show all events"""
        self.manager.clear_station_filter()
        self.refresh_display()
        self.update_status("Showing all events")
    
    def browse_sequence_path_click(self, sender, e):
        """Handle browse sequence path button click"""
        dialog = FolderBrowserDialog()
        dialog.SelectedPath = self.txt_sequence_path.Text
        if dialog.ShowDialog() == DialogResult.OK:
            self.txt_sequence_path.Text = dialog.SelectedPath
            config.set_sequence_path(dialog.SelectedPath)
    
    def show_configuration_click(self, sender, e):
        """Show configuration dialog"""
        config_dialog = ConfigurationDialog()
        config_dialog.ShowDialog()
    
    def show_template_manager_click(self, sender, e):
        """Show template manager"""
        template_dialog = TemplateSelectionDialog()
        template_dialog.ShowDialog()
    
    def generate_report_click(self, sender, e):
        """Generate Excel report for selected past events"""
        try:
            # Get selected events
            selected_events = self.events_grid.get_selected_events()
            
            if len(selected_events) == 0:
                MessageBox.Show(
                    "Please select events (check the boxes) to generate reports for.",
                    "No Events Selected",
                    MessageBoxButtons.OK,
                    MessageBoxIcon.Information
                )
                return
            
            # Filter for past events only
            now = datetime.utcnow()
            past_events = [e for e in selected_events if e.event_datetime and e.event_datetime < now]
            
            if len(past_events) == 0:
                MessageBox.Show(
                    "No past events selected. Reports can only be generated for events that have already occurred.\n\n" +
                    f"Selected events: {len(selected_events)}\nPast events: 0",
                    "No Past Events",
                    MessageBoxButtons.OK,
                    MessageBoxIcon.Warning
                )
                return
            
            if len(past_events) < len(selected_events):
                result = MessageBox.Show(
                    f"Only {len(past_events)} of {len(selected_events)} selected events have occurred.\n\n" +
                    "Generate reports for past events only?",
                    "Past Events Filter",
                    MessageBoxButtons.YesNo,
                    MessageBoxIcon.Question
                )
                if result != DialogResult.Yes:
                    return
            
            # Check if openpyxl is available
            try:
                import openpyxl
            except ImportError:
                MessageBox.Show(
                    "Excel reporting requires the 'openpyxl' library.\n\n" +
                    "To install, run in PowerShell:\n" +
                    "pip install openpyxl\n\n" +
                    "After installation, restart the application.",
                    "Missing Library",
                    MessageBoxButtons.OK,
                    MessageBoxIcon.Error
                )
                return
            
            # Generate reports
            self.lbl_status.Text = "Generating reports..."
            self.Refresh()
            
            success_count = 0
            report_folder = os.path.join(config.get_file_folder(), 'reports')
            if not os.path.exists(report_folder):
                os.makedirs(report_folder)
            
            for event in past_events:
                try:
                    report_path = self.generate_event_report(event, report_folder)
                    if report_path:
                        success_count += 1
                except Exception as ex:
                    print(f"Error generating report for {event.name}: {str(ex)}")
            
            self.lbl_status.Text = f"Generated {success_count} reports"
            
            MessageBox.Show(
                f"Successfully generated {success_count} report(s) for past events.\n\n" +
                f"Reports saved to: {report_folder}",
                "Reports Generated",
                MessageBoxButtons.OK,
                MessageBoxIcon.Information
            )
            
        except Exception as ex:
            MessageBox.Show(
                f"Error generating reports: {str(ex)}",
                "Error",
                MessageBoxButtons.OK,
                MessageBoxIcon.Error
            )
    
    def download_na_report_template(self):
        """Download North American Occultation Report Form template"""
        na_report_url = 'https://astrid-downloads.s3.amazonaws.com/downloads/NorthAmerica_AstReportForm_V5.6.12.xlsx'
        template_path = os.path.join(config.get_file_folder(), 'NorthAmerica_AstReportForm_V5.6.12.xlsx')
        
        if os.path.exists(template_path):
            return template_path
        
        try:
            import urllib.request
            self.lbl_status.Text = "Downloading NA Report Form template..."
            self.Refresh()
            urllib.request.urlretrieve(na_report_url, template_path)
            return template_path
        except Exception as ex:
            MessageBox.Show(
                f"Failed to download NA Report Form template:\n{str(ex)}\n\n" +
                f"Please download manually from:\n{na_report_url}\n\n" +
                f"Save to: {template_path}",
                "Template Download Failed",
                MessageBoxButtons.OK,
                MessageBoxIcon.Warning
            )
            return None
    
    def generate_event_report(self, event, report_folder):
        """Generate Excel report for a single event using NA Report Form template"""
        from openpyxl import load_workbook
        
        # Get template
        template_path = self.download_na_report_template()
        if not template_path:
            return None
        
        # Load template workbook
        try:
            wb = load_workbook(template_path)
            ws = wb['DATA']
            
            # Validate template
            if ws['G1'].value != 'Asteroid Occultation Report Form':
                raise ValueError('Invalid NA Report Form template')
        except Exception as ex:
            MessageBox.Show(
                f"Error loading template: {str(ex)}",
                "Template Error",
                MessageBoxButtons.OK,
                MessageBoxIcon.Error
            )
            return None
        
        # Cell mapping from astrid fillinnareport.py
        MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        
        cell_mapping = {
            'AstNum': 'E7',
            'AstName': 'K7',
            'EventYear': 'D5',
            'EventMonth': 'K5',
            'EventDay': 'P5',
            'StarCatalog': 'S7',
            'StarNumber': 'X7',
            'PredictedHours': 'Y5',
            'PredictedMinutes': 'AA5',
            'PredictedSeconds': 'AC5',
            'LatitudeFormat': 'E17',
            'LongitudeFormat': 'N17',
            'Latitude': 'E18',
            'LatitudeDir': 'J18',
            'Longitude': 'N18',
            'LongitudeDir': 'R18',
            'Elevation': 'V18',
            'ElevationUnits': 'W18',
            'ElevationDatum': 'AA18',
            'Timing': 'E22',
            'TimingDevice': 'E23',
            'Detector': 'E25',
            'OtherDetectorRelatedInfo': 'V25',
            'ObserverName': 'D9',
            'ObserverEmail': 'S9',
            'Aperture': 'E20',
            'ApertureUnits': 'H20',
            'FocalRatio': 'L20',
            'TelescopeType': 'T20',
            'StartedObservingHours': 'F31',
            'StartedObservingMins': 'H31',
            'StartedObservingSecs': 'J31',
            'StoppedObservingHours': 'F37',
            'StoppedObservingMins': 'H37',
            'StoppedObservingSecs': 'J37',
            'VideoFormat': 'L25',
            'ExposureIntegration': 'P25',
            'CommentLine3': 'D44',
        }
        
        # Fill in event data
        # Asteroid number and name
        if event.object_no:
            ws[cell_mapping['AstNum']] = event.object_no
        if event.object_name:
            ws[cell_mapping['AstName']] = event.object_name
        
        # Event date/time
        if event.event_datetime:
            ws[cell_mapping['EventYear']] = event.event_datetime.year
            ws[cell_mapping['EventMonth']] = MONTHS[event.event_datetime.month - 1]
            ws[cell_mapping['EventDay']] = event.event_datetime.day
            ws[cell_mapping['PredictedHours']] = event.event_datetime.hour
            ws[cell_mapping['PredictedMinutes']] = event.event_datetime.minute
            ws[cell_mapping['PredictedSeconds']] = event.event_datetime.second
        
        # Star catalog and number (parse from star_id)
        if event.star_name:
            star = event.star_name
            star_catalog = None
            star_number = None
            
            if star.startswith('TYC'):
                star_catalog = 'TYC       xxxx-xxxxx-x'
                star_number = star.replace('TYC ', '')
            elif star.startswith('HIP'):
                star_catalog = 'HIP  xxxxxx'
                star_number = star.replace('HIP ', '')
            elif star.startswith('UCAC2'):
                star_catalog = 'UCAC2        xxxxxxxx'
                star_number = star.replace('UCAC2 ', '')
            elif star.startswith('UCAC3'):
                star_catalog = 'UCAC3     xxx - xxxxxx'
                star_number = star.replace('UCAC3 ', '')
            elif star.startswith('UCAC4'):
                star_catalog = 'UCAC4     xxx - xxxxxx'
                star_number = star.replace('UCAC4 ', '')
            elif star.startswith('G'):
                star_catalog = 'G-coords hhmmss.s?ddmmss'
                star_number = star.replace('G', '')
            elif star.startswith('URAT1'):
                star_catalog = 'URAT1    xxx - xxxxxxx'
                star_number = star.replace('URAT1 ', '')
            elif star.startswith('1B'):
                star_catalog = '1B    xxx - xxxxxxx'
                star_number = star.replace('1B ', '')
            elif star.startswith('1N'):
                star_catalog = '1N    xxx - xxxxxxx'
                star_number = star.replace('1N ', '')
            
            if star_catalog and star_number:
                ws[cell_mapping['StarCatalog']] = star_catalog
                ws[cell_mapping['StarNumber']] = star_number
        
        # Observer location from config
        observer_lat = config.get_observer_latitude()
        observer_lon = config.get_observer_longitude()
        observer_elev = config.get_observer_elevation()
        
        if observer_lat != 0.0:
            ws[cell_mapping['LatitudeFormat']] = 'deg.ddddd'
            ws[cell_mapping['Latitude']] = '%0.5f' % abs(observer_lat)
            ws[cell_mapping['LatitudeDir']] = 'S' if observer_lat < 0 else 'N'
        
        if observer_lon != 0.0:
            ws[cell_mapping['LongitudeFormat']] = 'deg.ddddd'
            ws[cell_mapping['Longitude']] = '%0.5f' % abs(observer_lon)
            ws[cell_mapping['LongitudeDir']] = 'W' if observer_lon < 0 else 'E'
        
        if observer_elev != 0.0:
            ws[cell_mapping['Elevation']] = observer_elev
            ws[cell_mapping['ElevationUnits']] = 'm'
            ws[cell_mapping['ElevationDatum']] = 'WGS84'
        
        # Timing (SharpCap with GPS)
        ws[cell_mapping['Timing']] = 'GPS - other linking'
        ws[cell_mapping['TimingDevice']] = 'SharpCap'
        ws[cell_mapping['Detector']] = 'SharpCap'
        
        # Detector info
        ws[cell_mapping['OtherDetectorRelatedInfo']] = f'Exp {event.exposure_ms}ms'
        
        # Observer info from config
        observer_name = config.get_observer_name()
        observer_email = config.get_observer_email()
        if observer_name:
            ws[cell_mapping['ObserverName']] = observer_name
        if observer_email:
            ws[cell_mapping['ObserverEmail']] = observer_email
        
        # Telescope info from config
        aperture = config.get_telescope_aperture()
        focal_length = config.get_telescope_focal_length()
        telescope_type = config.get_telescope_type()
        
        if aperture > 0 and focal_length > 0:
            ws[cell_mapping['Aperture']] = aperture / 10.0  # Convert mm to cm
            ws[cell_mapping['ApertureUnits']] = 'cm'
            focal_ratio = focal_length / aperture
            ws[cell_mapping['FocalRatio']] = focal_ratio
        
        if telescope_type:
            ws[cell_mapping['TelescopeType']] = telescope_type
        
        # Recording times (event start/end converted to H:M:S)
        if event.start_time:
            ws[cell_mapping['StartedObservingHours']] = event.start_time.hour
            ws[cell_mapping['StartedObservingMins']] = event.start_time.minute
            ws[cell_mapping['StartedObservingSecs']] = event.start_time.second
        
        if event.end_time:
            ws[cell_mapping['StoppedObservingHours']] = event.end_time.hour
            ws[cell_mapping['StoppedObservingMins']] = event.end_time.minute
            ws[cell_mapping['StoppedObservingSecs']] = event.end_time.second
        
        # Video format and integration
        ws[cell_mapping['VideoFormat']] = 'SER'
        ws[cell_mapping['ExposureIntegration']] = 'Other'
        
        # Comment line
        ws[cell_mapping['CommentLine3']] = 'This report was pre-filled by Occultation Manager'
        
        # Generate filename (IOTA standard format: YYYYMMDD_asteroidnumber_asteroidname_observersurname_stationnumber_POS.xlsx)
        event_date = event.event_datetime.strftime('%Y%m%d') if event.event_datetime else 'unknown'
        clean_name = "".join(c for c in event.name if c.isalnum() or c in ('(', ')', ' ', '-', '_')).rstrip()
        filename = f"{event_date}_{clean_name}_Report.xlsx"
        report_path = os.path.join(report_folder, filename)
        
        # Save workbook
        wb.save(report_path)
        return report_path
    
    def show_about_click(self, sender, e):
        """Show about dialog"""
        MessageBox.Show("Occultation Manager for SharpCap\nVersion 1.0\n\nManages OccultWatcher Cloud events and generates SharpCap sequences.\n\nFeatures:\n- Download events from OW Cloud\n- Edit individual event exposures\n- Generate sequence files\n- Automatic sequence regeneration", 
                       "About", MessageBoxButtons.OK, MessageBoxIcon.Information)
    
    def exit_click(self, sender, e):
        """Exit application"""
        self.Close()
    
    def grid_selection_changed(self, sender, e):
        """Handle grid selection change"""
        if self.events_grid.SelectedRows.Count > 0:
            selected_row = self.events_grid.SelectedRows[0]
            if selected_row.Tag:
                event = selected_row.Tag
                details = f"Event: {event.event_name}\n"
                details += f"Star: {event.star_name}\n"
                details += f"Asteroid: {event.asteroid_name}\n"
                details += f"Time: {event.event_time_utc}\n"
                details += f"Magnitude: {event.magnitude:.1f}\n"
                details += f"Exposure: {event.exposure_ms}ms"
                if event.has_custom_exposure():
                    details += " (Custom)"
                details += f"\nDuration: {event.recording_duration}s"
                self.lbl_event_details.Text = details

###############################
# Sequence Generation Function
###############################

def save_occultation_sequence(occ, template_path="", sequence_path=None):
    """Format occultation data into readable report and save it"""
    if sequence_path is None:
        sequence_path = config.get_sequence_path()
    
    # Load template
    template_content = TemplateManager.load_template(template_path)
    if not template_content:
        print("No template file found!")
        return False
    
    # Handle both OccultationEvent objects and dictionaries
    if hasattr(occ, 'start_time_str'):  # It's an OccultationEvent object
        start_time = datetime.strptime(occ.start_time_str, '%Y-%m-%dT%H:%M:%S')
        clean_name = "".join(c for c in occ.name if c.isalnum() or c in ('(',')',' ', '-', '_')).rstrip()
        seq_name = start_time.strftime('%Y%m%d') + ' ' + clean_name + '.scs'
        
        occ_dict = {
            'object_name': occ.object_name,
            'event_time': occ.event_time,
            'start_time': occ.start_time_str,
            'goto_time': occ.goto_time_str,
            'recording_duration': occ.recording_duration,
            'star_mag': occ.star_mag,
            'comb_mag': occ.comb_mag,
            'mag_drop': occ.mag_drop,
            'event_uncertainty': occ.event_uncertainty,
            'ra': occ.ra,
            'dec': occ.dec,
            'asteroid_name': occ.object_name,
            'exposure': occ.get_exposure_seconds(),  # Use current exposure (custom or calculated)
            'name': occ.name,
            'station_name': occ.station_name
        }
    else:  # It's a dictionary (legacy format)
        start_time = datetime.strptime(occ['start_time'], '%Y-%m-%dT%H:%M:%S')
        clean_name = "".join(c for c in occ['name'] if c.isalnum() or c in ('(',')',' ', '-', '_')).rstrip()
        seq_name = start_time.strftime('%Y%m%d') + ' ' + clean_name + '.scs'
        occ_dict = occ
    
    try:
        # Ensure directory exists
        if not os.path.exists(sequence_path):
            os.makedirs(sequence_path, exist_ok=True)
        
        full_seq_path = os.path.join(sequence_path, seq_name)
        
        report = template_content.format(
            object_name=occ_dict.get('object_name', ''),
            event_time=occ_dict.get('event_time', ''),
            start_time=occ_dict.get('start_time', ''),
            goto_time=occ_dict.get('goto_time', ''),
            recording_duration=occ_dict.get('recording_duration', 0),
            star_mag=occ_dict.get('star_mag', 0),
            comb_mag=occ_dict.get('comb_mag', 0),
            mag_drop=occ_dict.get('mag_drop', 0),
            time_error=occ_dict.get('event_uncertainty', 0),
            ra=occ_dict.get('ra', 0),
            dec=occ_dict.get('dec', 0),
            asteroid_name=occ_dict.get('object_name', ''),
            exposure=occ_dict.get('exposure', 0)
        )
        
        with open(full_seq_path, 'w') as f:
            f.write(report)
        
        return True
        
    except Exception as e:
        print(f"Error creating sequence: {e}")
        return False

###############################
# Main Entry Point
###############################

if __name__ == "__main__":
    print(f"Configuration loaded from: {config.get_config_path()}")
    print(f"Working directory: {config.get_file_folder()}")
    
    # Validate configuration
    errors = config.validate_config()
    if errors:
        print("Configuration validation errors:")
        for error in errors:
            print(f"  - {error}")
        print("Please check your configuration settings.")
    
    try:
        print("Starting GUI mode...")
        app = OccultationManagerGUI()
        Application.EnableVisualStyles()
        Application.Run(app)
    except Exception as ex:
        print(f"GUI failed to start: {ex}")
        MessageBox.Show(f"Failed to start application: {ex}", "Startup Error", 
                       MessageBoxButtons.OK, MessageBoxIcon.Error)

